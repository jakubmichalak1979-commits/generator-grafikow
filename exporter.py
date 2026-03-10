import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import calendar
from datetime import date
import holidays
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors as rl_colors

def export_schedule(schedule_dict, year, month, filepath, location_name=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grafik Pracy"
    
    pl_holidays = holidays.Poland(years=year)
    month_names = ["", "STYCZEŃ", "LUTY", "MARZEC", "KWIECIEŃ", "MAJ", "CZERWIEC", 
                   "LIPIEC", "SIERPIEŃ", "WRZESIEŃ", "PAŹDZIERNIK", "LISTOPAD", "GRUDZIEŃ"]
    
    # Nagłówki Górne
    ws.merge_cells('A1:AI1')
    ws['A1'] = f"GRAFIK - OBIEKT {location_name.upper()}"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:AI2')
    ws['A2'] = f"MIESIĄC {month_names[month]} ROK {year}"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Nagłówki Tabeli (Wiersz 4)
    headers = ["Imię i Nazwisko"]
    num_days = calendar.monthrange(year, month)[1]
    for d in range(1, num_days + 1):
        headers.append(str(d))
        
    stats_headers = ["R", "P", "N", "W", "U", "CH", "S", "WE"]
    headers.extend(stats_headers)
    ws.append([]) # Pusty wiersz 3
    ws.append(headers)
    header_row = 4
    
    # Kolory czcionek
    colors = {
        'R': '00B050', 'P': '0070C0', 'N': '000000',
        'U': 'FF0000', 'CH': 'FF0000', 'W': 'FF0000'
    }
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        if 2 <= col <= num_days + 1:
            d = col - 1
            dt = date(year, month, d)
            if dt.weekday() == 6 or dt in pl_holidays:
                cell.fill = red_fill
            elif dt.weekday() == 5:
                cell.fill = green_fill
            
    # Dane
    for row_idx, (emp, days) in enumerate(schedule_dict.items(), start=header_row + 1):
        row_data = [emp]
        counts = {'R': 0, 'P': 0, 'N': 0, 'W': 0, 'U': 0, 'CH': 0, 'WE': 0}
        
        for d in range(1, num_days + 1):
            shift = days.get(d, '')
            if shift in ['R', 'P', 'N', 'W', 'U', 'CH']:
                counts[shift] += 1
            
            # Count WE/ŚW work
            dt = date(year, month, d)
            if (dt.weekday() >= 5 or dt in pl_holidays) and shift in ['R', 'P', 'N']:
                counts['WE'] += 1
            
            row_data.append(shift)
            
        s_sum = counts['R'] + counts['P'] + counts['N'] + counts['U'] + counts['CH']
        row_data.extend([counts['R'], counts['P'], counts['N'], counts['W'], counts['U'], counts['CH'], s_sum, counts['WE']])
        ws.append(row_data)
        
        for d in range(1, num_days + 1):
            cell = ws.cell(row=row_idx, column=1 + d)
            shift = cell.value
            dt = date(year, month, d)
            if shift in colors:
                cell.font = Font(color=colors[shift], bold=True)
            if dt.weekday() == 6 or dt in pl_holidays:
                cell.fill = red_fill
            elif dt.weekday() == 5:
                cell.fill = green_fill
            cell.alignment = Alignment(horizontal='center')
            
    wb.save(filepath)
    return filepath

def export_schedule_pdf(schedule_dict, year, month, filepath, location_name=""):
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                            rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=10)
    
    pl_holidays = holidays.Poland(years=year)
    num_days = calendar.monthrange(year, month)[1]
    month_names = ["", "STYCZEŃ", "LUTY", "MARZEC", "KWIECIEŃ", "MAJ", "CZERWIEC", 
                   "LIPIEC", "SIERPIEŃ", "WRZESIEŃ", "PAŹDZIERNIK", "LISTOPAD", "GRUDZIEŃ"]
    
    # Content
    elements = []
    from reportlab.lib.styles import getSampleStyleSheet
    styles = getSampleStyleSheet()
    title1 = f"<b>GRAFIK - OBIEKT {location_name.upper()}</b>"
    title2 = f"<b>MIESIĄC {month_names[month]} ROK {year}</b>"
    
    from reportlab.platypus import Paragraph, Spacer
    elements.append(Paragraph(title1, styles['Title']))
    elements.append(Paragraph(title2, styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    # Nagłówki Tabeli
    # R P N W U CH S WE (8 columns)
    headers = ["Imię i Naz."] + [str(d) for d in range(1, num_days + 1)] + ["R", "P", "N", "W", "U", "CH", "S", "WE"]
    data = [headers]
    
    for emp, days in schedule_dict.items():
        counts = {'R': 0, 'P': 0, 'N': 0, 'W': 0, 'U': 0, 'CH': 0, 'WE': 0}
        row = [emp[:10]]
        for d in range(1, num_days + 1):
            shift = days.get(d, '')
            if shift in ['R', 'P', 'N', 'W', 'U', 'CH']:
                counts[shift] += 1
            
            dt = date(year, month, d)
            if (dt.weekday() >= 5 or dt in pl_holidays) and shift in ['R', 'P', 'N']:
                counts['WE'] += 1
            
            row.append(shift)
        
        s_sum = counts['R'] + counts['P'] + counts['N'] + counts['U'] + counts['CH']
        row.extend([counts['R'], counts['P'], counts['N'], counts['W'], counts['U'], counts['CH'], s_sum, counts['WE']])
        data.append(row)
        
    # Widokość kolumn (Dynamiczna redukcja dla dużej ilości kolumn)
    # 1 (Name) + num_days + 8 (Stats)
    total_cols = 1 + num_days + 8
    # A4 landscape is ~842 points. 20 margin = 822 left.
    name_w = 55
    stat_w = 15
    day_w = (822 - name_w - (8 * stat_w)) / num_days
    col_widths = [name_w] + [day_w] * num_days + [stat_w] * 8
    
    t = Table(data, colWidths=col_widths)
    
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), rl_colors.lightgrey),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1, -1), 0.3, rl_colors.black),
    ])
    
    for row_idx, row in enumerate(data):
        if row_idx == 0:
            for d in range(1, num_days + 1):
                dt = date(year, month, d)
                if dt.weekday() == 6 or dt in pl_holidays:
                    style.add('BACKGROUND', (d, 0), (d, len(schedule_dict)), rl_colors.mistyrose)
                elif dt.weekday() == 5:
                    style.add('BACKGROUND', (d, 0), (d, len(schedule_dict)), rl_colors.honeydew)
        
        for col_idx, cell in enumerate(row):
            if cell == 'R': style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), rl_colors.green)
            elif cell == 'P': style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), rl_colors.blue)
            elif cell in ('U', 'CH', 'W'): style.add('TEXTCOLOR', (col_idx, row_idx), (col_idx, row_idx), rl_colors.red)
                
    t.setStyle(style)
    elements.append(t)
    doc.build(elements)
    return filepath
